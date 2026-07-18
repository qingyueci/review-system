from io import BytesIO
import math
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import SECTOR_COLORS

FONT_NAME = "微软雅黑"
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF0F8")
LIGHT_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BLOCK_SIDE = Side(style="medium", color="4472C4")


def _text(value) -> str:
    if isinstance(value, list):
        return "\n".join(str(item) for item in value if item not in (None, ""))
    return "" if value is None else str(value)


def _safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .") or "未知"


def _sector_color(sector: str) -> str:
    return next((color for keyword, color in SECTOR_COLORS.items() if keyword in sector), "E2E3E5")


def _style_cells(ws, row: int, start_col: int = 1, end_col: int = 5, *, fill=None, bold=False, color="000000", center=False) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name=FONT_NAME, size=10, bold=bold, color=color)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = LIGHT_BORDER
        if fill:
            cell.fill = fill


def _frame_block(ws, start_row: int, end_row: int) -> None:
    """用蓝色外框标出一个完整板块，内部保留浅色分隔线。"""
    for row in range(start_row, end_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.border = Border(
                left=BLOCK_SIDE if col == 1 else cell.border.left,
                right=BLOCK_SIDE if col == 5 else cell.border.right,
                top=BLOCK_SIDE if row == start_row else cell.border.top,
                bottom=BLOCK_SIDE if row == end_row else cell.border.bottom,
            )


def _text_width(value: str) -> int:
    """按 Excel 显示宽度估算字符长度，中文通常占两个英文字符宽度。"""
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in value)


def _available_width(ws, row: int, col: int) -> float:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sum(
                ws.column_dimensions[get_column_letter(index)].width or 13
                for index in range(merged.min_col, merged.max_col + 1)
            )
    return ws.column_dimensions[get_column_letter(col)].width or 13


def _fit_row_height(ws, row: int, minimum: float = 24, maximum: float = 120) -> None:
    """根据列宽、合并区域、中文长度和手动换行估算合适行高。"""
    max_lines = 1
    for col in range(1, 6):
        cell = ws.cell(row, col)
        if cell.value in (None, ""):
            continue
        width = max(6, int(_available_width(ws, row, col)))
        line_count = 0
        for line in str(cell.value).split("\n"):
            line_count += max(1, math.ceil(_text_width(line) / width))
        max_lines = max(max_lines, line_count)
    ws.row_dimensions[row].height = min(maximum, max(minimum, max_lines * 18 + 6))


def _add_section(ws, row: int, title: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, title)
    _style_cells(ws, row, fill=SECTION_FILL, bold=True, color="FFFFFF")
    ws.cell(row, 1).font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    ws.row_dimensions[row].height = 26
    return row + 1


def _add_label_row(ws, row: int, label: str, value, *, value_fill=None) -> int:
    ws.cell(row, 1, label)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row, 2, _text(value))
    _style_cells(ws, row)
    ws.cell(row, 1).fill = LABEL_FILL
    ws.cell(row, 1).font = Font(name=FONT_NAME, size=10, bold=True)
    if value_fill:
        for col in range(2, 6):
            ws.cell(row, col).fill = value_fill
    _fit_row_height(ws, row)
    return row + 1


def _write_first_boards(ws, row: int, data: dict) -> int:
    row = _add_section(ws, row, "一、首板复盘")
    for board in data["first_boards"]:
        if not isinstance(board, dict):
            continue
        block_start = row
        sector = _text(board.get("sector"))
        fill = PatternFill("solid", fgColor=_sector_color(sector))
        row = _add_label_row(ws, row, "板块", sector, value_fill=fill)

        stocks = [_text(stock) for stock in board.get("stocks", []) if _text(stock)]
        stock_groups = [stocks[index:index + 4] for index in range(0, len(stocks), 4)] or [[]]
        for group_index, group in enumerate(stock_groups):
            ws.cell(row, 1, "个股" if group_index == 0 else "")
            for offset in range(4):
                ws.cell(row, offset + 2, group[offset] if offset < len(group) else "")
            _style_cells(ws, row, center=True)
            ws.cell(row, 1).fill = LABEL_FILL
            ws.cell(row, 1).font = Font(name=FONT_NAME, size=10, bold=True)
            _fit_row_height(ws, row, minimum=26)
            row += 1

        row = _add_label_row(ws, row, "首封时间", board.get("first_seal_time"))
        row = _add_label_row(ws, row, "分析", board.get("analysis_points"))
        row = _add_label_row(ws, row, "板块预期", board.get("expectation"))
        _frame_block(ws, block_start, row - 1)
        row += 1

    if not data["first_boards"]:
        row = _add_label_row(ws, row, "板块", "暂无数据")
    row = _add_label_row(ws, row, "首板总结", data.get("first_board_summary"))
    return row + 1


def _write_ladders(ws, row: int, data: dict) -> int:
    row = _add_section(ws, row, "二、连板梯队")
    headers = ["板数", "个股", "晋级分析", "梯队思路", "梯队名称"]
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header)
    _style_cells(ws, row, fill=SUBHEADER_FILL, bold=True, center=True)
    row += 1

    ladders = sorted(
        (item for item in data["ladders"] if isinstance(item, dict)),
        key=lambda item: int(item.get("level") or 0),
        reverse=True,
    )
    for ladder in ladders:
        level = int(ladder.get("level") or 0)
        stocks = [stock for stock in ladder.get("stocks", []) if isinstance(stock, dict)]
        if not stocks:
            stocks = [{}]
        for stock in stocks:
            values = [
                level or "", _text(stock.get("name")), _text(stock.get("analysis")),
                _text(ladder.get("ladder_thought")), _text(ladder.get("level_name")),
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row, col, value)
            _style_cells(ws, row)
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if level >= 4:
                ws.cell(row, 1).fill = PatternFill("solid", fgColor="F8D7DA")
                ws.cell(row, 1).font = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")
            elif level == 2:
                ws.cell(row, 1).fill = PatternFill("solid", fgColor="E2E3E5")
            _fit_row_height(ws, row)
            row += 1
    if not ladders:
        ws.cell(row, 1, "暂无数据")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        _style_cells(ws, row)
        row += 1
    return row + 1


def _write_sentiment(ws, row: int, data: dict) -> int:
    row = _add_section(ws, row, "三、高标情绪")
    sentiment = data["sentiment"]
    row = _add_label_row(ws, row, "情绪描述", sentiment["high_sentiment"])
    row = _add_label_row(ws, row, "情绪标签", sentiment["mood_tag"])
    score = sentiment["mood_score"]
    score_fill = PatternFill("solid", fgColor="D4EDDA" if score >= 7 else "F8D7DA" if score <= 3 else "FFF3CD")
    return _add_label_row(ws, row, "强度分数", score, value_fill=score_fill) + 1


def _write_numbered_section(ws, row: int, title: str, label: str, values: list) -> int:
    row = _add_section(ws, row, title)
    ws.cell(row, 1, "序号")
    ws.cell(row, 2, label)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    _style_cells(ws, row, fill=SUBHEADER_FILL, bold=True, center=True)
    row += 1
    for number, value in enumerate(values, 1):
        ws.cell(row, 1, number)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row, 2, _text(value))
        _style_cells(ws, row)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        _fit_row_height(ws, row)
        row += 1
    if not values:
        ws.cell(row, 1, "暂无数据")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        _style_cells(ws, row)
        row += 1
    return row + 1


def _write_temperament(ws, row: int, data: dict) -> int:
    row = _add_section(ws, row, "六、气质股")
    headers = ["个股", "气质逻辑", "", "", "风险提示"]
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    _style_cells(ws, row, fill=SUBHEADER_FILL, bold=True, center=True)
    row += 1
    stocks = [item for item in data["temperament_stocks"] if isinstance(item, dict)]
    for item in stocks:
        ws.cell(row, 1, _text(item.get("name")))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, _text(item.get("logic")))
        ws.cell(row, 5, _text(item.get("risk")))
        _style_cells(ws, row)
        _fit_row_height(ws, row)
        row += 1
    if not stocks:
        ws.cell(row, 1, "暂无数据")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        _style_cells(ws, row)
        row += 1
    return row + 1


def generate_excel(data: dict) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "复盘汇总"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 14
    for column in ("B", "C", "D", "E"):
        ws.column_dimensions[column].width = 22

    ws.merge_cells("A1:E1")
    ws["A1"] = f"日期：{data['meta']['date']}"
    _style_cells(ws, 1, fill=PatternFill("solid", fgColor="1F4E78"), bold=True, color="FFFFFF", center=True)
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    title = data["meta"].get("title") or "复盘报告"
    author = data["meta"].get("author")
    ws["A2"] = f"{title}{'　作者：' + author if author else ''}"
    _style_cells(ws, 2, fill=PatternFill("solid", fgColor="DCE6F1"), center=True)
    ws.row_dimensions[2].height = 26

    row = 4
    row = _write_first_boards(ws, row, data)
    row = _write_ladders(ws, row, data)
    row = _write_sentiment(ws, row, data)
    row = _write_numbered_section(ws, row, "四、观察计划", "观察要点", data["observation_plan"])
    row = _write_numbered_section(ws, row, "五、竞价分析", "竞价逻辑", data["bidding_analysis"])
    row = _write_temperament(ws, row, data)
    row = _write_numbered_section(ws, row, "七、思考题", "问题内容", data["thinking_questions"])

    ws.print_area = f"A1:E{row - 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.oddFooter.center.text = "第 &P 页，共 &N 页"

    output = BytesIO()
    wb.save(output)
    author_name = _safe_filename(data["meta"]["author"])
    filename = f"复盘_{author_name}_{_safe_filename(data['meta']['date'])}.xlsx"
    return output.getvalue(), filename
