from io import BytesIO

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from review_app.excel import generate_excel
from review_app.preprocessing import preprocess_text
from review_app.ui import _preview
from review_app.validation import validate_data


def sample_data():
    return validate_data({
        "meta": {"date": "2026-07-14", "author": "测试作者", "title": "复盘"},
        "first_boards": [{"sector": "芯片", "stocks": ["甲公司", "乙公司"], "first_seal_time": "09:35", "analysis_points": ["资金回流"], "expectation": "观察持续性"}],
        "ladders": [{"level": 4, "stocks": [{"name": "甲公司", "analysis": "晋级"}], "ladder_thought": "高标观察"}],
        "sentiment": {"high_sentiment": ["情绪修复"], "mood_tag": "修复", "mood_score": 7},
        "observation_plan": ["观察量能"], "bidding_analysis": ["关注竞价承接"],
        "temperament_stocks": [{"name": "甲公司", "logic": "辨识度", "risk": "注意分歧"}],
        "thinking_questions": ["持续性如何？"],
    })


def test_preprocess_text():
    assert preprocess_text("  标题\n\n\n下载某某APP\n 正文") == "标题\n\n正文"


def test_generate_excel_uses_one_summary_sheet():
    content, filename = generate_excel(sample_data())
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["复盘汇总"]
    assert filename == "复盘_测试作者_2026-07-14.xlsx"
    sheet = workbook["复盘汇总"]
    assert sheet["A1"].value == "日期：2026-07-14"
    assert "A1:E1" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet["B6"].value == "甲公司"
    assert sheet["C6"].value == "乙公司"
    assert sheet["B15"].value == "甲公司"
    assert sheet["C15"].value == "晋级"


def test_first_board_groups_four_stocks_and_ladder_uses_one_stock_per_row():
    data = sample_data()
    data["first_boards"][0]["stocks"] = ["个股一", "个股二", "个股三", "个股四", "个股五", "个股六"]
    data["ladders"][0]["stocks"] = [
        {"name": "连板甲", "analysis": "分析甲"},
        {"name": "连板乙", "analysis": "分析乙"},
    ]
    content, _ = generate_excel(data)
    sheet = load_workbook(BytesIO(content))["复盘汇总"]
    assert [sheet.cell(6, col).value for col in range(2, 6)] == ["个股一", "个股二", "个股三", "个股四"]
    assert [sheet.cell(7, col).value for col in range(2, 4)] == ["个股五", "个股六"]
    assert [sheet["B16"].value, sheet["C16"].value] == ["连板甲", "分析甲"]
    assert [sheet["B17"].value, sheet["C17"].value] == ["连板乙", "分析乙"]


def test_each_sector_uses_an_independent_framed_block():
    data = sample_data()
    data["first_boards"].append({
        "sector": "医药",
        "stocks": ["医药甲"],
        "first_seal_time": "10:01",
        "analysis_points": ["资金轮动"],
        "expectation": "观察承接",
    })
    content, _ = generate_excel(data)
    sheet = load_workbook(BytesIO(content))["复盘汇总"]
    assert sheet["B5"].value == "芯片"
    assert sheet["B11"].value == "医药"
    assert sheet["B6"].value == "甲公司"
    assert sheet["B12"].value == "医药甲"
    assert sheet["A5"].border.top.style == "medium"
    assert sheet["A9"].border.bottom.style == "medium"
    assert sheet["A11"].border.top.style == "medium"
    assert sheet["A15"].border.bottom.style == "medium"


def test_content_rows_adjust_height_with_upper_limit():
    data = sample_data()
    data["first_boards"][0]["analysis_points"] = ["这是一段用于验证自动行高的长分析内容" * 12]
    content, _ = generate_excel(data)
    sheet = load_workbook(BytesIO(content))["复盘汇总"]
    assert sheet.row_dimensions[8].height > sheet.row_dimensions[7].height
    assert sheet.row_dimensions[8].height <= 120


def test_preview_labels_are_list_compatible():
    labels = list(_preview(sample_data()))
    assert labels == ["首板复盘", "连板梯队", "高标情绪", "观察计划", "竞价分析", "气质股", "思考题"]


def test_app_renders_seven_tabs_with_result():
    app = AppTest.from_file("app.py")
    app.run(timeout=10)
    app.session_state["review_data"] = sample_data()
    app.session_state["excel_content"] = b"test"
    app.session_state["excel_filename"] = "test.xlsx"
    app.session_state["analyzed_review_text"] = "测试复盘"
    app.run(timeout=10)
    assert not app.exception
    labels = [tab.label for tab in app.tabs]
    assert set(["今日复盘", "知识库", "刺大框架分析"]).issubset(labels)
    assert set(["首板复盘", "连板梯队", "思考题"]).issubset(labels)
